import openpyxl
from openpyxl.styles import PatternFill
from Bio.Seq import Seq
import random

# Define the length of the sequence
seq_length = 5000

# Define the nucleotides to choose from
nucleotides = ["A", "C", "G", "T"]

# Generate the random sequence
seq = "".join([random.choice(nucleotides) for _ in range(seq_length)])

# Function to split the sequence into chunks of length 'size'
def chunk_sequence(sequence, size):
    return [sequence[i:i+size] for i in range(0, len(sequence), size)]

# Function to export the sequence to an XLSX file
def export_sequence_to_xlsx(sequence, filename):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set cell colors based on nucleotide
    color_mapping = {
        'A': '00FF0000',  # Red
        'T': '0000FF00',  # Green
        'G': '000000FF',  # Blue
        'C': '00FFFF00',  # Yellow
    }

    # Split the sequence into chunks of 50 bases
    chunks = chunk_sequence(sequence, 50)

    # Write each base in a separate cell with color
    for row, chunk in enumerate(chunks, start=1):
        for col, base in enumerate(chunk, start=1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = base
            cell.fill = PatternFill(start_color=color_mapping.get(base, 'FFFFFFFF'),
                                    end_color=color_mapping.get(base, 'FFFFFFFF'),
                                    fill_type='solid')

            # Set column width to 2
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2

    # Save the workbook
    workbook.save(filename)

# Export the random sequence to XLSX file
export_sequence_to_xlsx(seq, "Sequence.xlsx")

# Read the XLSX file
workbook = openpyxl.load_workbook("Sequence.xlsx")
worksheet = workbook.active

# Extract the sequence from the XLSX file
extracted_seq = ""
for row in worksheet.iter_rows():
    for cell in row:
        value = cell.value
        if value is not None:
            extracted_seq += value

# Calculate complement sequence
complement_seq = Seq(extracted_seq).complement()

# Calculate reverse complement sequence
reverse_complement_seq = Seq(extracted_seq).reverse_complement()

# Export complement sequence to XLSX file
export_sequence_to_xlsx(str(complement_seq), "Complement.xlsx")

# Export reverse complement sequence to XLSX file
export_sequence_to_xlsx(str(reverse_complement_seq), "ReverseComplement.xlsx")


